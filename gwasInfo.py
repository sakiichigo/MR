# import libraries
import urllib.request
from bs4 import BeautifulSoup
import openpyxl as op
import socket
import requests
import random
import json
from hashlib import md5

#筛选过程，对获取的excel结果表填充人口、样本量、nsps、年份信息

# change excelUrl
excelUrl="C:/Users/user/Desktop/test1.xlsx"

# get population,sample_size,nsp,year
def getInfomation(id):
    url = 'https://gwas.mrcieu.ac.uk/datasets/' + id + '/'
    while(True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:',error,' ',url)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    pages = soup.find_all('th', class_='text-nowrap')
    population = "Nome"
    sample_size = "Nome"
    nsp = "Nome"
    year = "Nome"
    pmid = "Nome"
    consortium = "Nome"
    for i in range(0,len(pages),1):
        temp=pages[i]

        if (temp.text == "PMID"):
            pmid = temp.findNext("td").text
            continue
        if(temp.text=="Year"):
            year = temp.findNext("td").text
            continue
        if(temp.text=="Population"):
            population = temp.findNext("td").text
            continue
        if (temp.text == "Sample size"):
            sample_size = temp.findNext("td").text
            continue
        if (temp.text == "Number of SNPs"):
            nsp = temp.findNext("td").text
            continue
        if (temp.text == "Consortium"):
            consortium = temp.findNext("td").text
            continue
    result=[population,sample_size,nsp,year,pmid,consortium]
    return result

# baidu api translate
appid = '20231130001896707'
appkey = 'MkW4prMH1Nk2XI8dYdOY'
from_lang = 'en'
to_lang = 'zh'
endpoint = 'http://api.fanyi.baidu.com'
path = '/api/trans/vip/translate'
url = endpoint + path

# Generate salt and sign
def make_md5(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()

def baidu_api(query, from_lang, to_lang):
    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()
    return result["trans_result"][0]['dst']

# check element
def checkElement(check,row,column):
    if(sheet.cell(row=row, column=column).value==None):
        sheet.insert_cols(column)
    sheet.cell(row=row, column=column, value=check)

# modify xlsx
wb = op.load_workbook(excelUrl)
for i in range(0,len(wb.worksheets),1):
    sheet = wb.worksheets[i]
    #if(sheet.cell(row=1, column=17).value!=None):
     #   continue
    checkElement('population', 1, 17)
    checkElement('sample_size', 1, 18)
    checkElement('number_of_snps', 1, 19)
    checkElement('year', 1, 20)
    checkElement('exposure_zh', 1, 21)
    checkElement('outcome_zh', 1, 22)
    checkElement('pmid', 1, 23)
    checkElement('Consortium', 1, 24)
    for j in range(2,sheet.max_row+1,1):
        exposure=sheet.cell(row=j, column=1).value
        outcome=sheet.cell(row=j, column=2).value
        outcome_en = sheet.cell(row=j, column=3).value
        exposure_en = sheet.cell(row=j, column=4).value
        if(exposure==None or outcome==None):
            continue
        if(exposure!=sheet.cell(row=j-1, column=1).value):
            exposure_zh=baidu_api(exposure_en.split("||")[0].replace("_"," ").replace("-"," "),from_lang, to_lang)
            inf_exposure = getInfomation(exposure)
        if(outcome!=sheet.cell(row=j-1, column=2).value):
            outcome_zh = baidu_api(outcome_en.split("||")[0].replace("_"," ").replace("-"," "), from_lang, to_lang)
            inf_outcome = getInfomation(outcome)
        population_exposure=inf_exposure[0]
        population_outcome=inf_outcome[0]
        print("exposure|outcome population :"+population_exposure+"|"+population_outcome)
        #population,sample_size,nsp,year,pmid,consortium
        if(population_exposure==population_outcome):
            sheet.cell(row=j, column=17,value=population_exposure)
        sheet.cell(row=j, column=18, value=str(inf_exposure[1])+"|"+inf_outcome[1])
        sheet.cell(row=j, column=19, value=str(inf_exposure[2])+"|"+inf_outcome[2])
        sheet.cell(row=j, column=20, value=str(inf_exposure[3]) + "|" + inf_outcome[3])
        sheet.cell(row=j, column=21, value=exposure_zh)
        sheet.cell(row=j, column=22, value=outcome_zh)
        sheet.cell(row=j, column=23, value=str(inf_exposure[4])+"|"+inf_outcome[4])
        sheet.cell(row=j, column=24, value=str(inf_exposure[5])+"|"+inf_outcome[5])
    wb.save(excelUrl)
    print(wb.worksheets[i].title+" saved")