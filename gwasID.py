# import libraries
import urllib.request
from bs4 import BeautifulSoup
import openpyxl as op
import os

# change trait and path
trait="taste"
path="C:/Users/user/Desktop/idp/"
isGwasInfo=True

# params
url = 'https://gwas.mrcieu.ac.uk/datasets/?gwas_id__icontains=&year__iexact=&trait__icontains='+trait+'&consortium__icontains='
rows = []

# get GWAS id
def getRows(url,sheet,row):
    while (True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:', error)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    table = soup.find('table')
    results = table.find_all('td', class_="text-nowrap")
    if (len(results) == 0):
        print("no available data")
    else:
        if(isGwasInfo):
            for result in results:
                data = result.string
                print(data)
                info = getInfomation(data)
                sheet.cell(row=row, column=1, value=data)
                for j in range(0, len(info), 1):
                    sheet.cell(row=row, column=j+2, value=info[j])
                row += 1
        else:
            for result in results:
                data = result.string
                print(data)
                rows.append(data)
            # export
            with open(path + trait + ".txt", "w") as file:
                for line in rows:
                    file.write(line + "\n")


def getInfomation(id):
    url = 'https://gwas.mrcieu.ac.uk/datasets/' + id + '/'
    while(True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:',error,' ',url)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    pages = soup.find_all('th', class_='text-nowrap')
    population = ""
    sample_size = ""
    nsp = ""
    year = ""
    pmid = ""
    consortium = ""
    for i in range(0, len(pages), 1):
        temp = pages[i]
        if (temp.text == "PMID"):
            pmid = temp.findNext("td").text
            continue
        if (temp.text == "Year"):
            year = temp.findNext("td").text
            continue
        if (temp.text == "Population"):
            population = temp.findNext("td").text
            continue
        if (temp.text == "Sample size"):
            sample_size = temp.findNext("td").text
            continue
        if (temp.text == "Number of SNPs"):
            nsp = temp.findNext("td").text
            continue
        if (temp.text == "Consortium"):
            consortium = temp.findNext("td").text
            continue
    result = [population, sample_size, nsp, year, pmid, consortium]
    return result

# fill rows
while (True):
    try:
        page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
    except Exception as error:
        print('Error:', error)
    else:
        break
soup = BeautifulSoup(page, 'html.parser')
pages = soup.find_all('a', class_='page-link')
if (os.path.exists(path + trait + '.xlsx')):
    wb = op.load_workbook(filename=trait + '.xlsx')
    sheet = wb.worksheets[0]
    row = sheet.max_row + 1
else:
    wb = op.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value="id")
    sheet.cell(row=1, column=2, value="population")
    sheet.cell(row=1, column=3, value="sample_size")
    sheet.cell(row=1, column=4, value="number_of_snps")
    sheet.cell(row=1, column=5, value="year")
    row = 2
if (len(pages) <= 2):
    getRows(url, sheet, row)
else:
    maxPage = pages[len(pages) - 2].contents[0].replace("\n", "").replace(" ", "")
    maxPage = int(maxPage)
    for pageCout in range(1, maxPage + 1, 1):
        urlpage = url + '&page=' + str(pageCout)
        getRows(urlpage, sheet, row)
        row = sheet.max_row + 1
wb.save(path + trait + '.xlsx')
wb.close()
print(trait + " created")
